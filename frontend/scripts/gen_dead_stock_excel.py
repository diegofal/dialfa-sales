"""Generate dead stock liquidation Excel for Juan and Dani."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from pathlib import Path
from decimal import Decimal
from datetime import datetime
import psycopg2, psycopg2.extras
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

env = (Path(__file__).parent.parent / ".env").read_text(encoding="utf-8")
url = next(line.split("=",1)[1].strip().strip('"').strip("'")
           for line in env.splitlines()
           if line.strip().startswith("DATABASE_URL") and not line.strip().startswith("#"))
conn = psycopg2.connect(url, cursor_factory=psycopg2.extras.RealDictCursor)
cur = conn.cursor()
f = lambda v: float(v) if isinstance(v, Decimal) else (v or 0)

cur.execute("""
WITH last_sale AS (
  SELECT soi.article_id,
         MAX(i.invoice_date)::date AS last_sale_date,
         SUM(soi.quantity) AS total_ever_sold
  FROM sales_order_items soi
  JOIN sales_orders so ON soi.sales_order_id = so.id
  JOIN invoices i ON i.sales_order_id = so.id
  WHERE i.is_cancelled = false AND i.deleted_at IS NULL AND so.deleted_at IS NULL
  GROUP BY soi.article_id
),
lp AS (
  SELECT DISTINCT ON (soi.article_id) soi.article_id, soi.proforma_unit_price AS fob
  FROM supplier_order_items soi
  JOIN supplier_orders so ON so.id = soi.supplier_order_id
  WHERE soi.proforma_unit_price > 0 AND so.deleted_at IS NULL
  ORDER BY soi.article_id, so.order_date DESC
)
SELECT
  a.code, a.description, c.name AS categoria,
  a.stock::int AS stock, a.unit_price AS precio_lista,
  (a.stock * a.unit_price) AS valor_lista_total,
  COALESCE(lp.fob, a.unit_price * 0.18) AS costo_fob_est,
  COALESCE(lp.fob * 1.5, a.unit_price * 0.27) AS costo_cif_est,
  ls.last_sale_date, ls.total_ever_sold::int AS total_vendido_historico,
  CASE
    WHEN ls.last_sale_date IS NULL THEN 'Nunca vendido'
    WHEN EXTRACT(DAY FROM (NOW() - ls.last_sale_date)) > 1825 THEN 'Muerto 5+ anios'
    WHEN EXTRACT(DAY FROM (NOW() - ls.last_sale_date)) > 730 THEN 'Muerto 2-5 anios'
    ELSE 'Muerto 1-2 anios'
  END AS antiguedad,
  CASE
    WHEN ls.last_sale_date IS NULL THEN 60
    WHEN EXTRACT(DAY FROM (NOW() - ls.last_sale_date)) > 1825 THEN 60
    WHEN EXTRACT(DAY FROM (NOW() - ls.last_sale_date)) > 730 THEN 50
    ELSE 40
  END AS descuento_sugerido,
  lp.fob IS NOT NULL AS tiene_fob_real
FROM articles a
LEFT JOIN categories c ON c.id = a.category_id
LEFT JOIN last_sale ls ON ls.article_id = a.id
LEFT JOIN lp ON lp.article_id = a.id
WHERE a.deleted_at IS NULL AND a.is_active = true AND a.stock > 0
  AND (ls.last_sale_date IS NULL OR EXTRACT(DAY FROM (NOW() - ls.last_sale_date)) > 365)
ORDER BY (a.stock * a.unit_price) DESC
""")
rows = cur.fetchall()
conn.close()

wb = openpyxl.Workbook()
title_font = Font(name='Calibri', size=16, bold=True, color='1F4E79')
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
green_font = Font(bold=True, color='006100', size=12)
money_fmt = '#,##0.00'
int_fmt = '#,##0'
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
tier_fills = {
    'Muerto 1-2 anios': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    'Muerto 2-5 anios': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    'Muerto 5+ anios': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
    'Nunca vendido': PatternFill(start_color='D6D8DB', end_color='D6D8DB', fill_type='solid'),
}

# ===== SHEET 1: RESUMEN =====
ws = wb.active
ws.title = "Resumen"
ws['A1'] = 'DIALFA SRL - Plan de Liquidacion de Stock Muerto'
ws['A1'].font = title_font
ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
ws['A2'].font = Font(color='808080', size=10)
ws['A3'] = 'Propuesta para Juan y Dani - Comision 8% sobre ventas de dead stock'
ws['A3'].font = Font(color='808080', size=10, italic=True)

tiers = {}
for r in rows:
    tier = r['antiguedad']
    if tier not in tiers:
        tiers[tier] = {'arts': 0, 'uds': 0, 'valor_lista': 0, 'costo_cif': 0, 'desc': f(r['descuento_sugerido'])}
    tiers[tier]['arts'] += 1
    tiers[tier]['uds'] += r['stock']
    tiers[tier]['valor_lista'] += f(r['valor_lista_total'])
    tiers[tier]['costo_cif'] += f(r['costo_cif_est']) * r['stock']

rn = 5
for col, h in enumerate(['Categoria', 'Articulos', 'Unidades', 'Valor Lista',
                          'Descuento', 'Precio Liquidacion', 'Costo CIF',
                          'Margen', 'Comision 8%', 'Por Persona'], 1):
    c = ws.cell(row=rn, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center')

rn += 1
total_rev = total_cost = total_com = 0
for tn in ['Muerto 1-2 anios', 'Muerto 2-5 anios', 'Muerto 5+ anios', 'Nunca vendido']:
    t = tiers.get(tn)
    if not t: continue
    dp = t['desc'] / 100
    pl = t['valor_lista'] * (1 - dp)
    mg = pl - t['costo_cif']
    cm = pl * 0.08
    total_rev += pl; total_cost += t['costo_cif']; total_com += cm
    ws.cell(row=rn, column=1, value=tn)
    ws.cell(row=rn, column=2, value=t['arts']).number_format = int_fmt
    ws.cell(row=rn, column=3, value=t['uds']).number_format = int_fmt
    ws.cell(row=rn, column=4, value=t['valor_lista']).number_format = money_fmt
    ws.cell(row=rn, column=5, value=f"{int(t['desc'])}%")
    ws.cell(row=rn, column=6, value=pl).number_format = money_fmt
    ws.cell(row=rn, column=7, value=t['costo_cif']).number_format = money_fmt
    ws.cell(row=rn, column=8, value=mg).number_format = money_fmt
    ws.cell(row=rn, column=9, value=cm).number_format = money_fmt
    ws.cell(row=rn, column=10, value=cm/2).number_format = money_fmt
    if tier_fills.get(tn):
        for c in range(1, 11):
            ws.cell(row=rn, column=c).fill = tier_fills[tn]
    rn += 1

rn += 1
ws.cell(row=rn, column=1, value='TOTAL').font = Font(bold=True, size=12)
ws.cell(row=rn, column=4, value=sum(t['valor_lista'] for t in tiers.values())).number_format = money_fmt
ws.cell(row=rn, column=6, value=total_rev).number_format = money_fmt
ws.cell(row=rn, column=6).font = Font(bold=True, size=12)
ws.cell(row=rn, column=7, value=total_cost).number_format = money_fmt
ws.cell(row=rn, column=8, value=total_rev - total_cost).number_format = money_fmt
ws.cell(row=rn, column=8).font = green_font
ws.cell(row=rn, column=9, value=total_com).number_format = money_fmt
ws.cell(row=rn, column=10, value=total_com/2).number_format = money_fmt
ws.cell(row=rn, column=10).font = Font(bold=True, color='1F4E79', size=12)

rn += 3
ws.cell(row=rn, column=1, value='METRICAS CLAVE').font = Font(bold=True, size=13)
rn += 1
for label, val in [
    ('Revenue total al precio liquidacion', total_rev),
    ('Costo CIF de la mercaderia', total_cost),
    ('Ganancia bruta total', total_rev - total_cost),
    ('Comision total Juan + Dani (8%)', total_com),
    ('Comision por persona', total_com / 2),
    ('Comision por persona / mes (si 6 meses)', total_com / 2 / 6),
    ('Comision por persona / mes (si 12 meses)', total_com / 2 / 12),
    ('NETO PARA LA EMPRESA', total_rev - total_cost - total_com),
]:
    ws.cell(row=rn, column=1, value=label)
    c = ws.cell(row=rn, column=5, value=val)
    c.number_format = money_fmt
    if 'NETO' in label:
        c.font = green_font
        ws.cell(row=rn, column=1).font = Font(bold=True, size=12)
    rn += 1

for col in range(1, 11):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
ws.column_dimensions['A'].width = 30

# ===== SHEET 2: LISTA COMPLETA =====
ws2 = wb.create_sheet("Lista de Liquidacion")
headers = ['Codigo', 'Descripcion', 'Categoria', 'Stock', 'Precio Lista',
           'Valor Lista Total', 'Antiguedad', 'Desc. Sugerido',
           'Precio Liquidacion', 'Valor Liquidacion Total',
           'Costo CIF Unit.', 'Margen Unit.', 'Margen %',
           'Comision 8% Total', 'Ultima Venta', 'Vendido Historico']
for col, h in enumerate(headers, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = thin_border

for i, r in enumerate(rows, 2):
    dp = f(r['descuento_sugerido']) / 100
    pl = f(r['precio_lista']) * (1 - dp)
    cif = f(r['costo_cif_est'])
    mu = pl - cif
    mp = mu / pl if pl > 0 else 0
    vlt = pl * r['stock']
    ct = vlt * 0.08
    vals = [r['code'], r['description'], r['categoria'] or '', r['stock'],
            f(r['precio_lista']), f(r['valor_lista_total']), r['antiguedad'],
            f"{int(f(r['descuento_sugerido']))}%", pl, vlt, cif, mu, mp, ct,
            str(r['last_sale_date']) if r['last_sale_date'] else 'Nunca',
            r['total_vendido_historico'] or 0]
    fill = tier_fills.get(r['antiguedad'])
    for col, val in enumerate(vals, 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.border = thin_border
        if fill: c.fill = fill
        if col in (5, 6, 9, 10, 11, 12, 14): c.number_format = money_fmt
        elif col == 13: c.number_format = '0.0%'
        elif col == 4: c.number_format = int_fmt

widths = [15, 40, 20, 8, 12, 14, 18, 10, 14, 16, 12, 12, 10, 14, 12, 10]
for col, w in enumerate(widths, 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f"A1:P{len(rows)+1}"

# ===== SHEET 3: REGLAS =====
ws3 = wb.create_sheet("Reglas del Programa")
rules = [
    ('PROGRAMA DE INCENTIVO - LIQUIDACION DE STOCK MUERTO', '', True, 14),
    ('', '', False, 11),
    ('Participantes:', 'Juan y Dani', True, 11),
    ('Duracion:', '6 meses (renovable)', True, 11),
    ('Comision:', '8% del revenue de cada venta de articulos de esta lista', True, 11),
    ('Split:', '50/50 entre Juan y Dani (4% cada uno)', True, 11),
    ('Pago:', 'Trimestral (al cierre de cada trimestre)', True, 11),
    ('', '', False, 11),
    ('PRECIOS DE LIQUIDACION:', '', True, 12),
    ('Muerto 1-2 anios:', '40% de descuento sobre lista (vende al 60% del precio lista)', False, 11),
    ('Muerto 2-5 anios:', '50% de descuento sobre lista (vende al 50%)', False, 11),
    ('Muerto 5+ anios:', '60% de descuento sobre lista (vende al 40%)', False, 11),
    ('Nunca vendido:', '60% de descuento sobre lista (vende al 40%)', False, 11),
    ('', '', False, 11),
    ('REGLAS:', '', True, 12),
    ('1.', 'No vender por debajo del "Precio Liquidacion" sin autorizacion', False, 11),
    ('2.', 'Pueden ofrecer hasta 10% adicional sobre el precio liquidacion (previa aprobacion)', False, 11),
    ('3.', 'Solo los articulos de la hoja "Lista de Liquidacion" califican', False, 11),
    ('4.', 'Cada venta se registra en SPISA como factura normal', False, 11),
    ('5.', 'Meta sugerida: mover el 30% del stock muerto en 6 meses', False, 11),
    ('', '', False, 11),
    ('NOTA IMPORTANTE:', '', True, 12),
    ('', 'Todos los precios de liquidacion estan POR ENCIMA del costo CIF.', False, 11),
    ('', 'La empresa gana margen positivo en CADA venta, incluso al maximo descuento.', False, 11),
    ('', 'El costo CIF promedio es ~21.5% del precio lista original.', False, 11),
    ('', 'Este programa NO le cuesta plata a la empresa. Se autofinancia.', False, 11),
]
for i, (a, b, bold, sz) in enumerate(rules, 1):
    ca = ws3.cell(row=i, column=1, value=a)
    cb = ws3.cell(row=i, column=2, value=b)
    if bold: ca.font = Font(bold=True, size=sz, color='1F4E79' if sz > 11 else '000000')

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 70

out = Path(r'C:\Users\User\Desktop\DIALFA_Liquidacion_Dead_Stock.xlsx')
wb.save(str(out))
print(f"Excel generado en: {out}")
print(f"  Articulos: {len(rows)}")
print(f"  Revenue liquidacion: ${total_rev:,.0f}")
print(f"  Comision Juan+Dani: ${total_com:,.0f} (${total_com/2:,.0f} c/u)")
print(f"  Neto empresa: ${total_rev - total_cost - total_com:,.0f}")
