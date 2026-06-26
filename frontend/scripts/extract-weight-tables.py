# -*- coding: utf-8 -*-
"""
Extracts weight reference tables from the 3 Desktop Excel files into a single
normalized JSON file (scripts/output/weight-tables.json).

Each emitted record: { source, block, type, series, thickness, size, weight }
- `type` uses DB-style raw type strings so the TS MatchingKeyNormalizer maps them
  to the same canonical key as the DB `articles.type` column.
- `thickness` is expanded to every DB thickness the block's weight applies to.
"""
import json, re, unicodedata, os
import openpyxl

DESKTOP = os.path.expandvars(r"%USERPROFILE%\Desktop")
OUT = os.path.join(os.path.dirname(__file__), "output", "weight-tables.json")

def norm_header(s):
    if s is None:
        return ""
    # Drop degree / ordinal marks BEFORE NFKD (º and ª decompose to o/a otherwise)
    s = re.sub(r"[°ºª�]", " ", str(s))
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.upper().replace(".", " ")
    s = re.sub(r"[^A-Z0-9 /]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def is_size(v):
    if v is None:
        return False
    s = str(v).strip().strip('"').strip()
    return bool(re.match(r'^\d', s)) and ('"' in str(v) or "X" in s.upper() or re.match(r'^[\d\s/]+"?$', str(v).strip()))

records = []

def add(source, block, typ, series, thicknesses, size, weight):
    if weight is None or not isinstance(weight, (int, float)) or weight <= 0:
        return
    size = str(size).strip()
    for thk in thicknesses:
        records.append({
            "source": source, "block": block, "type": typ,
            "series": series, "thickness": thk, "size": size, "weight": float(weight),
        })

# ---------------------------------------------------------------- ACCESORIOS
# butt-weld fittings, series = None. Header text -> (db_type, [thicknesses])
ACC_MAP = {
    "CODOS 90 R L STD":   ("90D LR ELBOW", ["STD"]),
    "CODOS 90 R L E P":   ("90D LR ELBOW", ["XS"]),
    "CODOS 45 STD":       ("45D LR ELBOW", ["STD"]),
    "CODOS 45 E P":       ("45D LR ELBOW", ["XS"]),
    "CODOS 90 R C STD":   ("90D SR ELBOW", ["STD"]),
    "TEE STD":            ("TEE", ["STD"]),
    "TEE E P":            ("TEE", ["XS"]),
    "CASQUETES STD":      ("CAP", ["STD"]),
    "CAS E P":            ("CAP", ["XS"]),
    "TEE DE RED STD":     ("RED. TEE", ["STD"]),
    "RED CON STD":        ("CON. RED.", ["STD"]),
    "RED CON E P":        ("CON. RED.", ["XS"]),
    "RED EXEN STD":       ("EXC. RED.", ["STD"]),
}

wb = openpyxl.load_workbook(os.path.join(DESKTOP, "PESO DE ACCESORIOS.xlsx"), data_only=True)
ws = wb["ACCESORIOS"]
cur = None
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    c = ws.cell(r, 3).value
    h = norm_header(a)
    if h in ACC_MAP:
        cur = ACC_MAP[h]
        continue
    if cur and a is not None and isinstance(c, (int, float)) and str(a).strip().upper() not in ("DIAMETRO",):
        typ, thks = cur
        add("ACCESORIOS", h, typ, None, thks, a, c)

# ---------------------------------------------------------------- BRIDAS
# flanges. Header text -> (db_type, series, [thicknesses])
BRI_MAP = {
    "S O R F S 150":       ("SORF", 150, ["STD"]),
    "W N R F S 150":       ("W.N.R.F.", 150, ["40", "80", "160", "STD"]),
    "CIEGAS S 150":        ("BLIND", 150, ["STD"]),
    "ROSCADAS":            ("THREADED", 150, ["STD"]),
    "BRIDAS SORF S 300":   ("SORF", 300, ["STD"]),
    "BLIND S 300":         ("BLIND", 300, ["STD"]),
    "W N S 300 SCH 40":    ("W.N.R.F.", 300, ["40"]),
    "W N S 300 SCH 80":    ("W.N.R.F.", 300, ["80"]),
    "SORF S 600":          ("SORF", 600, ["STD"]),
    "BLIND S 600":         ("BLIND", 600, ["STD"]),
    "W N S 600 SCH 40":    ("W.N.R.F.", 600, ["40"]),
    "W N S 600 SCH 80":    ("W.N.R.F.", 600, ["80"]),
}
wb = openpyxl.load_workbook(os.path.join(DESKTOP, "PLANILLA PARA PEDIR BRIDAS.xlsx"), data_only=True)
ws = wb["BRIDAS"]
cur = None
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    c = ws.cell(r, 3).value
    h = norm_header(a)
    if h in BRI_MAP:
        cur = BRI_MAP[h]
        continue
    if cur and a is not None and isinstance(c, (int, float)) and str(a).strip().upper() not in ("DIAMETRO", "DIAMTRO"):
        typ, series, thks = cur
        add("BRIDAS", h, typ, series, thks, a, c)

# ---------------------------------------------------------------- FORJADO
# Forged A-105. Layout: TYPE header, then SERIES header (S-2000/3000/6000),
# then a column-header row labelling connection types (BSPT/NPT/SW) each
# followed by a PESO column. Each conn type -> different DB type.
FORJ_TYPE = {
    "CODOS 90": "90D LR THREADED {C} ELBOW",
    "CODOS 45": "45D LR THREADED {C} ELBOW",   # SW -> 45D S.W. ELBOW handled below
    "TEE": "TEE {C}",
    "UNION DOBLE": "UNION {C}",
    "CUPLAS": "COUPLING {C}",
}
def forj_type(block, conn):
    if block == "CODOS 45" and conn == "SW":
        return "45D S.W. ELBOW"
    if block == "CODOS 90" and conn == "SW":
        return "90D LR THREADED SW ELBOW"
    if block == "TEE" and conn == "SW":
        return "TEE S.W."
    tmpl = FORJ_TYPE.get(block)
    if not tmpl:
        return None
    return tmpl.replace("{C}", conn)

def parse_conn(label):
    l = norm_header(label)
    if l == "SW":
        return "SW"
    if l == "BSPT":
        return "BSPT"
    if l == "NPT":
        return "NPT"
    return None

wb = openpyxl.load_workbook(os.path.join(DESKTOP, "PLANILLA PARA PEDIDO DE FORJADO.xlsx.xlsx"), data_only=True)
ws = wb["Hoja1"]
cur_block = None
cur_series = None
col_conn = {}   # column index -> connection type, weight in col+1
for r in range(1, ws.max_row + 1):
    vals = [ws.cell(r, col).value for col in range(1, ws.max_column + 1)]
    a = vals[0]
    h = norm_header(a)
    # block header (CODOS 90, CODOS 45, TEE, UNION DOBLE, CUPLAS, ...)
    if h in FORJ_TYPE:
        cur_block = h
        cur_series = None
        col_conn = {}
        continue
    # series header (S 2000 / S 3000 / S 6000) anywhere in the row
    sm = None
    for v in vals:
        m = re.search(r"S\s*[- ]?\s*(2000|3000|6000)", norm_header(v))
        if m:
            sm = int(m.group(1)); break
    if sm:
        cur_series = sm
        col_conn = {}
        continue
    # column header row: cells containing BSPT/NPT/SW define conn columns
    conns_here = {i: parse_conn(v) for i, v in enumerate(vals) if parse_conn(v)}
    if conns_here:
        col_conn = conns_here   # weight is the column immediately to the right
        continue
    # data row: col A is a size, weights in (conn_col + 1)
    if cur_block and cur_series and col_conn and a is not None and is_size(a):
        for ci, conn in col_conn.items():
            w = vals[ci + 1] if ci + 1 < len(vals) else None
            typ = forj_type(cur_block, conn)
            if typ and isinstance(w, (int, float)):
                add("FORJADO", f"{cur_block} S-{cur_series}", typ, cur_series, ["STD"], a, w)

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(records, f, ensure_ascii=False, indent=0)

from collections import Counter
print(f"Extracted {len(records)} weight records -> {OUT}")
bysrc = Counter(r["source"] for r in records)
for k, v in bysrc.items():
    print(f"  {k}: {v}")
